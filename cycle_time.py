"""
Cycle Time Report — CAECS, GTM, DTAL, DTK
==========================================
Requirements:
    pip install requests python-dateutil openpyxl

Environment variables required:
    JIRA_EMAIL         helder.araujo@outsystems.com
    JIRA_API_TOKEN     Atlassian API token
    SLACK_WEBHOOK_URL  Slack incoming webhook (optional)

Methodology:
  - Only tickets with Story Points > 0 OR Fix Version filled are included
  - Cycle time = working days in ACTIVE statuses only
  - Weekends + Portuguese public holidays excluded
  - Blocked AND Test Blocked both excluded from cycle time, tracked separately
  - Ready / To Do / Backlog etc. excluded (inactive)
  - Epics and Discarded tickets are excluded from all metrics
"""

import os
import math
import requests
import concurrent.futures
from calendar import monthrange
from datetime import date, datetime, timezone, timedelta
from collections import defaultdict
from dateutil import parser as dp

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────
JIRA_BASE    = "https://outsystemsrd.atlassian.net/rest/api/3"
JIRA_BROWSE  = "https://outsystemsrd.atlassian.net/browse"
JIRA_EMAIL   = os.environ["JIRA_EMAIL"]
JIRA_TOKEN   = os.environ["JIRA_API_TOKEN"]
SLACK_URL    = os.environ.get("SLACK_WEBHOOK_URL")

PROJECTS  = ["CAECS", "GTM", "DTAL", "DTK"]
TEAM_DEVS = {"CAECS": 4, "GTM": 2, "DTAL": 1, "DTK": 2}

# Issue types excluded from all metrics
EXCLUDED_ISSUE_TYPES = {"epic"}

# Status classification
INACTIVE_STATUSES = {
    "to do", "open", "backlog", "new",
    "selected for development", "ready",
}
# Both "blocked" AND "test blocked" excluded from cycle time, tracked separately
BLOCKED_STATUSES = {"blocked", "test blocked"}
DONE_STATUSES    = {"done", "closed", "resolved"}

AUTH    = (JIRA_EMAIL, JIRA_TOKEN)
HEADERS = {"Accept": "application/json"}

# ── Portuguese Public Holidays ─────────────────────────────────────────────────

def easter_date(year):
    """Meeus/Jones/Butcher algorithm — returns UTC datetime of Easter Sunday."""
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4;    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4;    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day   = (h + l - 7 * m + 114) % 31 + 1
    return datetime(year, month, day, tzinfo=timezone.utc)


_holiday_cache = {}

def pt_holidays(year):
    """Returns a set of 'YYYY-MM-DD' strings for PT national holidays."""
    if year in _holiday_cache:
        return _holiday_cache[year]
    easter      = easter_date(year)
    good_friday = easter - timedelta(days=2)
    corpus      = easter + timedelta(days=60)
    fmt = lambda d: d.strftime("%Y-%m-%d")
    holidays = {
        f"{year}-01-01",  # New Year's Day
        f"{year}-04-25",  # Freedom Day
        f"{year}-05-01",  # Labour Day
        f"{year}-06-10",  # Portugal Day
        f"{year}-08-15",  # Assumption of Mary
        f"{year}-10-05",  # Republic Day
        f"{year}-11-01",  # All Saints' Day
        f"{year}-12-01",  # Restoration of Independence
        f"{year}-12-08",  # Immaculate Conception
        f"{year}-12-25",  # Christmas
        fmt(good_friday),
        fmt(corpus),
    }
    _holiday_cache[year] = holidays
    return holidays


def is_working_day(dt):
    """True if dt (UTC datetime) is a working day (Mon–Fri, not a PT holiday)."""
    if dt.weekday() >= 5:
        return False
    return dt.strftime("%Y-%m-%d") not in pt_holidays(dt.year)


def working_days_between(start_dt, end_dt):
    """
    Fractional working days between two UTC datetimes.
    Iterates day-by-day, clamping the overlap per day to [start_dt, end_dt].
    """
    if end_dt <= start_dt:
        return 0.0
    DAY_SECS = 86_400.0
    total    = 0.0
    cursor   = datetime(start_dt.year, start_dt.month, start_dt.day, tzinfo=timezone.utc)
    while cursor < end_dt:
        if is_working_day(cursor):
            day_end = cursor + timedelta(days=1)
            overlap = (min(day_end, end_dt) - max(cursor, start_dt)).total_seconds()
            total  += overlap / DAY_SECS
        cursor += timedelta(days=1)
    return total


def working_weeks_in_range(start_str, end_str):
    """Working days in [start, end] inclusive ÷ 5."""
    start = datetime.strptime(start_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    end   = datetime.strptime(end_str,   "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(hours=23, minutes=59, seconds=59)
    days  = 0
    cur   = start
    while cur <= end:
        if is_working_day(cur):
            days += 1
        cur += timedelta(days=1)
    return days / 5.0


# ── Jira API helpers ───────────────────────────────────────────────────────────

def jira_get(path, params=None):
    r = requests.get(
        f"{JIRA_BASE}/{path}",
        params=params, auth=AUTH, headers=HEADERS, timeout=30
    )
    r.raise_for_status()
    return r.json()


def get_all_issues(jql):
    """
    Paginate using search/jql endpoint.

    FILTERING RULE: Only include tickets where:
      - Story Points > 0 (customfield_10016 or customfield_10004), OR
      - fixVersions is non-empty
    """
    issues, page_token, page = [], None, 1
    print(f"  JQL: {jql}")
    while True:
        params = {
            "jql":        jql,
            "maxResults": 100,
            "fields":     "summary,project,created,resolutiondate,issuetype,"
                          "fixVersions,"
                          "customfield_10004,customfield_10016,"   # story points
                          "customfield_16256,"                     # AC Failed
                          "customfield_21332",                     # AI Dev Assist
        }
        if page_token:
            params["nextPageToken"] = page_token
        data  = jira_get("search/jql", params)
        batch = data.get("issues", [])
        issues.extend(batch)
        print(f"  Page {page}: {len(batch)} issues (total so far: {len(issues)})")
        if data.get("isLast", True) or not batch:
            break
        page_token = data.get("nextPageToken")
        if not page_token:
            break
        page += 1

    # ── Apply SP > 0 OR Fix Version filter ────────────────────────────────────
    before = len(issues)
    filtered = []
    for i in issues:
        sp = (i["fields"].get("customfield_10016")
           or i["fields"].get("customfield_10004"))
        has_sp  = sp is not None and float(sp) > 0
        fv      = i["fields"].get("fixVersions") or []
        has_fv  = isinstance(fv, list) and len(fv) > 0
        if has_sp or has_fv:
            filtered.append(i)

    excluded = before - len(filtered)
    if excluded > 0:
        print(f"  ⚙️  Filtered out {excluded} ticket(s) with no SP and no Fix Version "
              f"→ {len(filtered)} remaining")
    return filtered


def get_changelog(issue_key):
    data = jira_get(f"issue/{issue_key}", {
        "expand": "changelog",
        "fields": "status,customfield_10004,customfield_10016,"
                  "fixVersions,customfield_16256,customfield_21332",
    })
    fields    = data.get("fields", {})
    histories = data.get("changelog", {}).get("histories", [])

    story_points = (fields.get("customfield_10016")
                 or fields.get("customfield_10004"))

    ac_failed = False
    ac_val    = fields.get("customfield_16256")
    if isinstance(ac_val, dict):
        ac_failed = ac_val.get("value", "").lower() == "yes"
    elif isinstance(ac_val, str):
        ac_failed = ac_val.lower() == "yes"

    ai_field = fields.get("customfield_21332") or []
    if isinstance(ai_field, list):
        ai_tools = [v.get("value", str(v)) if isinstance(v, dict) else str(v)
                    for v in ai_field]
    else:
        ai_tools = []

    fix_versions = [v.get("name", "") for v in (fields.get("fixVersions") or [])]

    return {
        "histories":    histories,
        "story_points": story_points,
        "ac_failed":    ac_failed,
        "ai_tools":     ai_tools,
        "fix_versions": fix_versions,
    }


# ── Open Bugs ──────────────────────────────────────────────────────────────────

def get_open_bugs():
    """
    Fetch all open Bugs and Defects across the 4 projects.
    Excludes: Done, Discarded, Ready for Production.
    """
    jql = (
        'project in (CAECS, GTM, DTAL, DTK) '
        'AND issuetype in (Bug, Defect) '
        'AND status not in (Done, Discarded, "Ready for Production")'
    )
    issues, page_token = [], None
    while True:
        params = {
            "jql":        jql,
            "maxResults": 100,
            "fields":     "project,priority,status",
        }
        if page_token:
            params["nextPageToken"] = page_token
        data  = jira_get("search/jql", params)
        batch = data.get("issues", [])
        issues.extend(batch)
        if data.get("isLast", True) or not batch:
            break
        page_token = data.get("nextPageToken")
        if not page_token:
            break

    by_project  = {p: 0 for p in PROJECTS}
    high_urgent = 0
    blocked     = 0

    for i in issues:
        proj     = i["fields"]["project"]["key"]
        priority = (i["fields"].get("priority") or {}).get("name", "").lower()
        status   = i["fields"]["status"]["name"].lower()
        if proj in by_project:
            by_project[proj] += 1
        if priority in ("high", "urgent"):
            high_urgent += 1
        if status in BLOCKED_STATUSES:
            blocked += 1

    return {
        "total":       len(issues),
        "by_project":  by_project,
        "high_urgent": high_urgent,
        "blocked":     blocked,
    }


# ── Cycle Time Calculation ─────────────────────────────────────────────────────

def classify_status(name):
    n = name.lower().strip()
    if n in DONE_STATUSES:     return "done"
    if n in BLOCKED_STATUSES:  return "blocked"   # includes "test blocked"
    if n in INACTIVE_STATUSES: return "inactive"
    return "active"


def compute_cycle_time(histories):
    """
    Walk status transitions in chronological order.
    Returns (cycle_days, blocked_days, in_progress_days, status_log).

    Active      = counted towards cycle time
    In Progress = subset of Active, tracked separately for Efficiency metric
    Blocked     = "Blocked" OR "Test Blocked" — excluded from cycle, tracked separately
    Inactive    = To Do, Ready, Backlog, etc. — excluded entirely
    """
    transitions = []
    for h in histories:
        for item in h.get("items", []):
            if item["field"] == "status":
                transitions.append({
                    "date": dp.parse(h["created"]).astimezone(timezone.utc),
                    "from": (item.get("fromString") or "").strip(),
                    "to":   (item.get("toString")   or "").strip(),
                })
    transitions.sort(key=lambda x: x["date"])

    cycle_wd = blocked_wd = in_progress_wd = 0.0
    active_from = blocked_from = None
    status_log  = []

    for t in transitions:
        to_type  = classify_status(t["to"])
        frm_type = classify_status(t["from"])

        if to_type == "blocked":
            # Leaving active → entering blocked
            if active_from:
                d = working_days_between(active_from, t["date"])
                cycle_wd += d
                if t["from"].lower().strip() == "in progress":
                    in_progress_wd += d
                status_log.append({
                    "from": t["from"], "to": t["to"],
                    "counted": "✅ Active",
                    "days": round(d, 2),
                    "entered": active_from.strftime("%Y-%m-%d"),
                    "exited":  t["date"].strftime("%Y-%m-%d"),
                })
                active_from = None
            if not blocked_from:
                blocked_from = t["date"]

        elif frm_type == "blocked":
            # Leaving blocked
            if blocked_from:
                d = working_days_between(blocked_from, t["date"])
                blocked_wd += d
                status_log.append({
                    "from": t["from"], "to": t["to"],
                    "counted": "🚫 Blocked",
                    "days": round(d, 2),
                    "entered": blocked_from.strftime("%Y-%m-%d"),
                    "exited":  t["date"].strftime("%Y-%m-%d"),
                })
                blocked_from = None
            if to_type == "active":
                active_from = t["date"]

        elif to_type == "done":
            if active_from:
                d = working_days_between(active_from, t["date"])
                cycle_wd += d
                if t["from"].lower().strip() == "in progress":
                    in_progress_wd += d
                status_log.append({
                    "from": t["from"], "to": t["to"],
                    "counted": "✅ Active",
                    "days": round(d, 2),
                    "entered": active_from.strftime("%Y-%m-%d"),
                    "exited":  t["date"].strftime("%Y-%m-%d"),
                })
                active_from = None
            if blocked_from:
                d = working_days_between(blocked_from, t["date"])
                blocked_wd += d
                status_log.append({
                    "from": t["from"], "to": t["to"],
                    "counted": "🚫 Blocked",
                    "days": round(d, 2),
                    "entered": blocked_from.strftime("%Y-%m-%d"),
                    "exited":  t["date"].strftime("%Y-%m-%d"),
                })
                blocked_from = None

        elif to_type == "active" and not active_from and not blocked_from:
            active_from = t["date"]

        elif to_type == "inactive" and active_from:
            d = working_days_between(active_from, t["date"])
            cycle_wd += d
            if t["from"].lower().strip() == "in progress":
                in_progress_wd += d
            status_log.append({
                "from": t["from"], "to": t["to"],
                "counted": "✅ Active",
                "days": round(d, 2),
                "entered": active_from.strftime("%Y-%m-%d"),
                "exited":  t["date"].strftime("%Y-%m-%d"),
            })
            active_from = None

    return round(cycle_wd, 1), round(blocked_wd, 1), round(in_progress_wd, 1), status_log


# ── Excel Export ───────────────────────────────────────────────────────────────

C_HEADER = "1E2433"; C_PROJ = "2D3748"; C_ALT = "F7F8FC"; C_WHITE = "FFFFFF"
C_SUMMARY = "EBF4FF"

def hf(sz=11):  return Font(name="Arial", bold=True, color="FFFFFF", size=sz)
def bf(sz=10):  return Font(name="Arial", size=sz)
def lf():       return Font(name="Arial", size=10, color="4A6CF7", underline="single")
def bdr():
    s = Side(style="thin", color="D0D4E8")
    return Border(left=s, right=s, top=s, bottom=s)
def ca(): return Alignment(horizontal="center", vertical="center")
def la(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def fl(c): return PatternFill("solid", start_color=c, fgColor=c)


def export_excel(results, tickets, year, month, weeks):
    label = datetime(year, month, 1).strftime("%B %Y")
    wb    = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Summary"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"📊 Cycle Time Report — {label}  ({weeks:.1f} working weeks)"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fl(C_HEADER); ws["A1"].alignment = ca()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = ("Working days only · Ready, Blocked & Test Blocked excluded · "
                "weekends & 🇵🇹 PT holidays excluded · "
                "only tickets with SP > 0 or Fix Version")
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="94A3B8")
    ws["A2"].fill = fl(C_HEADER); ws["A2"].alignment = ca()
    ws.row_dimensions[2].height = 16

    ws.append([])  # row 3 blank

    hdrs = ["Project", "Tickets", "AVG Cycle (d)", "AVG Blocked (d)",
            "Throughput (SP/wk)", "Rework Rate", "AI Assist %", "Efficiency %", "Devs"]
    ws.append(hdrs)  # row 4
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=col)
        c.font = hf(); c.fill = fl(C_PROJ); c.alignment = ca(); c.border = bdr()
    ws.row_dimensions[4].height = 22

    for idx, p in enumerate(results):
        row = 5 + idx
        f   = fl(C_ALT if idx % 2 == 0 else C_WHITE)
        vals = [
            p["project"],
            p["count"],
            p["avg"],
            p["avg_blocked"],
            "N/A (SP<50%)" if p["sp_na"] else (p["throughput_sp"] if p["throughput_sp"] is not None else "—"),
            f"{p['rework_rate']}%" if p["rework_rate"] is not None else "—",
            f"{p['ai_rate']}%"     if p["ai_rate"]     is not None else "—",
            f"{p['efficiency']}%"  if p.get("efficiency") is not None else "—",
            TEAM_DEVS.get(p["project"], "?"),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = bf(); c.fill = f
            c.alignment = la() if col == 1 else ca()
            c.border = bdr()

    ws.column_dimensions["A"].width = 10
    for col, w in zip("BCDEFGHI", [9, 15, 16, 18, 12, 12, 13, 6]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ── Sheet 2: All Tickets ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("🎫 All Tickets")
    ws2.merge_cells("A1:J1")
    ws2["A1"] = (f"All Tickets — {label} ({len(tickets)} total) · "
                 "Working days, weekends & PT holidays excluded · SP > 0 or Fix Version only")
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = fl(C_HEADER); ws2["A1"].alignment = ca()
    ws2.row_dimensions[1].height = 26

    hdrs2 = ["Jira Key", "Project", "Summary", "SP", "Fix Version",
             "Cycle (wd)", "In Prog (wd)", "Blocked (wd)", "AC Failed", "AI Assist", "How counted"]
    ws2.append(hdrs2)
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=2, column=col)
        c.font = hf(); c.fill = fl(C_PROJ)
        c.alignment = la() if col in (3, 10) else ca(); c.border = bdr()
    ws2.row_dimensions[2].height = 20

    for idx, t in enumerate(sorted(tickets, key=lambda x: (x["project"], -x["cycle"])), 1):
        row = idx + 2
        f   = fl(C_ALT if idx % 2 == 0 else C_WHITE)

        kc = ws2.cell(row=row, column=1, value=t["key"])
        kc.hyperlink = f"{JIRA_BROWSE}/{t['key']}"
        kc.font = lf(); kc.fill = f; kc.alignment = ca(); kc.border = bdr()

        fix_ver_str = ", ".join(t.get("fix_versions") or []) or "—"
        for col, val in enumerate([
            t["project"], t["summary"],
            t.get("story_points", "—") or "—",
            fix_ver_str,
            t["cycle"],
            t.get("in_progress", "—"),
            t["blocked"],
            "🔁 Yes" if t.get("ac_failed") else "—",
            f"🤖 {', '.join(t.get('ai_tools', []))}" if t.get("ai_tools") else "—",
        ], 2):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = bf(); c.fill = f
            c.alignment = la() if col in (3,) else ca(); c.border = bdr()

        # How counted column
        if t.get("status_log"):
            parts = [f"{s['from']}→{s['to']} ({s['counted']}: {s['days']}d)"
                     for s in t["status_log"] if s["days"] > 0]
            how = " | ".join(parts) if parts else "No active time logged"
        else:
            how = f"Active: {t['cycle']}wd"
        hc = ws2.cell(row=row, column=11, value=how)
        hc.font = Font(name="Arial", size=9, color="475569")
        hc.fill = f; hc.alignment = la(); hc.border = bdr()

    ws2.column_dimensions["A"].width = 13
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["D"].width = 5
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 11
    ws2.column_dimensions["G"].width = 13
    ws2.column_dimensions["H"].width = 13
    ws2.column_dimensions["I"].width = 10
    ws2.column_dimensions["J"].width = 20
    ws2.column_dimensions["K"].width = 80
    ws2.freeze_panes = "A3"

    # ── One sheet per project ─────────────────────────────────────────────────
    for proj in PROJECTS:
        items = sorted([t for t in tickets if t["project"] == proj],
                       key=lambda x: -x["cycle"])
        ws_p  = wb.create_sheet(proj)
        ws_p.merge_cells("A1:H1")
        ws_p["A1"] = f"{proj} — {label} ({len(items)} tickets · SP > 0 or Fix Version)"
        ws_p["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws_p["A1"].fill = fl(C_HEADER); ws_p["A1"].alignment = ca()
        ws_p.row_dimensions[1].height = 26

        hdrs3 = ["Jira Key", "Summary", "SP", "Fix Version",
                 "Cycle (wd)", "Blocked (wd)", "AC Failed", "How counted"]
        ws_p.append(hdrs3)
        for col, h in enumerate(hdrs3, 1):
            c = ws_p.cell(row=2, column=col)
            c.font = hf(); c.fill = fl(C_PROJ)
            c.alignment = la() if col in (2, 8) else ca(); c.border = bdr()
        ws_p.row_dimensions[2].height = 20

        for idx, t in enumerate(items, 1):
            row = idx + 2
            f   = fl(C_ALT if idx % 2 == 0 else C_WHITE)

            kc2 = ws_p.cell(row=row, column=1, value=t["key"])
            kc2.hyperlink = f"{JIRA_BROWSE}/{t['key']}"
            kc2.font = lf(); kc2.fill = f; kc2.alignment = ca(); kc2.border = bdr()

            fix_ver_str = ", ".join(t.get("fix_versions") or []) or "—"
            for col, val in enumerate([
                t["summary"],
                t.get("story_points", "—") or "—",
                fix_ver_str,
                t["cycle"],
                t["blocked"],
                "🔁 Yes" if t.get("ac_failed") else "—",
            ], 2):
                c = ws_p.cell(row=row, column=col, value=val)
                color = ("22863A" if t["cycle"] < 10 else
                         "E36209" if t["cycle"] < 30 else "D73A49") if col == 5 else None
                c.font = Font(name="Arial", bold=(col == 5), size=10,
                              color=color if color else "000000")
                c.fill = f
                c.alignment = la() if col == 2 else ca(); c.border = bdr()

            if t.get("status_log"):
                parts2 = [f"{s['from']}→{s['to']} ({s['days']}wd)"
                          for s in t["status_log"] if s["days"] > 0]
                how2 = " | ".join(parts2) if parts2 else "No active time"
            else:
                how2 = f"{t['cycle']}wd"
            hc2 = ws_p.cell(row=row, column=8, value=how2)
            hc2.font = Font(name="Arial", size=9, color="475569")
            hc2.fill = f; hc2.alignment = la(); hc2.border = bdr()

        # AVG row
        ar = len(items) + 3
        ws_p[f"A{ar}"] = "AVG Cycle Time (wd)"
        ws_p[f"A{ar}"].font = Font(name="Arial", bold=True, size=11)
        ws_p[f"A{ar}"].fill = fl(C_SUMMARY); ws_p[f"A{ar}"].border = bdr()
        ws_p[f"E{ar}"] = f"=AVERAGE(E3:E{ar-1})"
        ws_p[f"E{ar}"].font = Font(name="Arial", bold=True, size=12, color="4A6CF7")
        ws_p[f"E{ar}"].fill = fl(C_SUMMARY); ws_p[f"E{ar}"].alignment = ca()
        ws_p[f"E{ar}"].border = bdr(); ws_p[f"E{ar}"].number_format = "0.0"

        ws_p.column_dimensions["A"].width = 13
        ws_p.column_dimensions["B"].width = 55
        ws_p.column_dimensions["C"].width = 5
        ws_p.column_dimensions["D"].width = 18
        ws_p.column_dimensions["E"].width = 12
        ws_p.column_dimensions["F"].width = 13
        ws_p.column_dimensions["G"].width = 10
        ws_p.column_dimensions["H"].width = 80
        ws_p.freeze_panes = "A3"

    filename = f"cycle_time_report_{year}_{month:02d}.xlsx"
    wb.save(filename)
    print(f"\n📁 Excel saved: {filename}")
    return filename


# ── Confluence Upload ──────────────────────────────────────────────────────────

CONFLUENCE_PAGE_ID = "6314983433"
CONFLUENCE_BASE    = "https://outsystemsrd.atlassian.net/wiki"

def upload_to_confluence(filepath):
    import base64
    filename = os.path.basename(filepath)
    mime     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    url      = f"{CONFLUENCE_BASE}/rest/api/content/{CONFLUENCE_PAGE_ID}/child/attachment"
    creds    = base64.b64encode(f"{JIRA_EMAIL}:{JIRA_TOKEN}".encode()).decode()
    headers  = {"Authorization": f"Basic {creds}", "X-Atlassian-Token": "nocheck"}

    print(f"   Uploading {filename}...")
    with open(filepath, "rb") as fh:
        file_data = fh.read()

    check    = requests.get(url, headers={**headers, "Accept": "application/json"},
                            params={"filename": filename}, timeout=30)
    existing = check.json().get("results", []) if check.status_code == 200 else []

    if existing:
        att_id  = existing[0]["id"]
        upd_url = (f"{CONFLUENCE_BASE}/rest/api/content/{CONFLUENCE_PAGE_ID}"
                   f"/child/attachment/{att_id}/data")
        r      = requests.post(upd_url, headers=headers,
                               files={"file": (filename, file_data, mime)}, timeout=60)
        action = "Updated"
    else:
        r      = requests.post(url, headers=headers,
                               files={"file": (filename, file_data, mime)}, timeout=60)
        action = "Uploaded"

    print(f"   HTTP {r.status_code}")
    if r.status_code in (200, 201):
        page_url = f"{CONFLUENCE_BASE}/spaces/DTHOME/pages/{CONFLUENCE_PAGE_ID}"
        print(f"   ✅ {action} to Confluence: {page_url}")
        return page_url
    print(f"   ⚠️  Confluence upload failed: {r.status_code} — {r.text[:400]}")
    return None


# ── Slack ──────────────────────────────────────────────────────────────────────

Q4_GOALS = {
    "lead_time_wd":        5.0,
    "throughput_growth":   20,
    "rework_rate":         10,
    "efficiency_min":      30,
    "efficiency_max":      40,
    "ai_assist":           50,
    "open_bugs":           5,
}

def _goal_indicator(value, goal, lower_is_better=True):
    if value is None or goal is None:
        return ""
    on_goal = value <= goal if lower_is_better else value >= goal
    return " ✅" if on_goal else " ⚠️"


def post_to_slack(report, confluence_url=None):
    if not SLACK_URL:
        print("\n⚠️  SLACK_WEBHOOK_URL not set — skipping Slack post")
        return

    def weighted_avg(projects):
        valid = [p for p in projects if p["avg"] is not None and p["count"] > 0]
        n = sum(p["count"] for p in valid)
        return f"{sum(p['avg']*p['count'] for p in valid)/n:.1f}" if n else "N/A"

    lines = [
        f"*📊 Monthly Cycle Time Report — {report['label']}*",
        f"_Projects: {', '.join(PROJECTS)}_",
        f"_Working days only · Weekends & PT holidays excluded · "
        f"Ready/Blocked/Test Blocked excluded · SP > 0 or Fix Version only_",
        "",
    ]

    for p in report["projects"]:
        avg     = p["avg"]
        avg_str = f"{avg:.1f}wd" if avg is not None else "N/A"
        goal_lt = Q4_GOALS["lead_time_wd"]
        lt_ind  = _goal_indicator(avg, goal_lt, lower_is_better=True)
        lead_time_line = f"Lead Time: *{avg_str}*{lt_ind} _(goal: <{goal_lt}wd)_"

        blocked_line = ""
        if p["avg_blocked"] > 0:
            blocked_line = f"\n        Blocked: {p['avg_blocked']:.1f}wd avg"

        if p["sp_na"]:
            tput_line = "Throughput: N/A (insufficient Story Points data)"
        elif p["throughput_sp"] is not None:
            tput_line = (f"Throughput: *{p['throughput_sp']:.1f} SP/week* · "
                         f"{p['throughput_sp_dev']:.1f} SP/dev/week")
        else:
            tput_line = "Throughput: N/A"

        rw     = p["rework_rate"]
        rw_str = f"{rw}%" if rw is not None else "N/A"
        rw_ind = _goal_indicator(rw, Q4_GOALS["rework_rate"], lower_is_better=True)
        rework_line = (f"Rework Rate: *{rw_str}*{rw_ind} "
                       f"_(goal: <{Q4_GOALS['rework_rate']}%)_ (Stories & Tasks only)")

        eff     = p.get("efficiency")
        eff_str = f"{eff}%" if eff is not None else "N/A"
        if eff is not None:
            on_goal    = Q4_GOALS["efficiency_min"] <= eff <= Q4_GOALS["efficiency_max"]
            eff_ind    = " ✅" if on_goal else " ⚠️"
        else:
            eff_ind = ""
        eff_line = (f"Efficiency: *{eff_str}*{eff_ind} "
                    f"_(goal: {Q4_GOALS['efficiency_min']}%–{Q4_GOALS['efficiency_max']}%)_")

        ai     = p["ai_rate"]
        ai_str = f"{ai}%" if ai is not None else "N/A"
        ai_ind = _goal_indicator(ai, Q4_GOALS["ai_assist"], lower_is_better=False)
        ai_line = f"AI Assisted Code: *{ai_str}*{ai_ind} _(goal: {Q4_GOALS['ai_assist']}%)_"

        lines += [
            f"*── {p['project']} ({p['count']} tickets) ──*",
            f"        {lead_time_line}{blocked_line}",
            f"        {tput_line}",
            f"        {rework_line}",
            f"        {eff_line}",
            f"        {ai_line}",
            "",
        ]

    bugs       = report.get("open_bugs", {})
    bugs_total = bugs.get("total", "N/A")
    bugs_goal  = Q4_GOALS["open_bugs"]
    if isinstance(bugs_total, int):
        bugs_ind    = " ✅" if bugs_total <= bugs_goal else " ⚠️"
        by_proj_str = "  ·  ".join(
            f"{proj}: {bugs['by_project'].get(proj, 0)}" for proj in PROJECTS
        )
        hi_str = f"  ·  High/Urgent: {bugs.get('high_urgent', 0)}" if bugs.get("high_urgent") else ""
        bl_str = f"  ·  Blocked: {bugs.get('blocked', 0)}"          if bugs.get("blocked")     else ""
        lines += [
            "*── Open Bugs & Defects ──*",
            f"        Total: *{bugs_total}*{bugs_ind} _(goal: <{bugs_goal})_",
            f"        {by_proj_str}{hi_str}{bl_str}",
            "",
        ]

    ov      = weighted_avg(report["projects"])
    goal_lt = Q4_GOALS["lead_time_wd"]
    try:
        ov_ind = " ✅" if float(ov) <= goal_lt else " ⚠️"
    except ValueError:
        ov_ind = ""

    lines += [
        "*── Overall ──*",
        f"        Lead Time AVG: *{ov}wd*{ov_ind} _(goal: <{goal_lt}wd)_ "
        f"across {report['total']} tickets · {report['weeks']:.1f} working weeks",
        "",
        "_Cycle time = active working days · SP > 0 or Fix Version · "
        "In Progress and equivalent statuses_",
    ]

    if confluence_url:
        lines.append(f"\n📎 *Full breakdown (Excel):* <{confluence_url}|Download from Confluence>")

    r = requests.post(SLACK_URL, json={"text": "\n".join(lines)}, timeout=10)
    r.raise_for_status()
    print("\n✅ Posted to Slack #dtge_v2mom_sync")


# ── Report ─────────────────────────────────────────────────────────────────────

def run_report(year, month):
    _, last_day = monthrange(year, month)
    start = f"{year}-{month:02d}-01"
    end   = f"{year}-{month:02d}-{last_day}"
    label = datetime(year, month, 1).strftime("%B %Y")
    weeks = working_weeks_in_range(start, end)

    print(f"\n📊 Cycle Time Report — {label}  ({weeks:.1f} working weeks)")
    print("=" * 60)
    print("⚙️  Filter: SP > 0 OR Fix Version filled · Epics & Discarded excluded")
    print("=" * 60)

    jql = (
        f'project in ({", ".join(PROJECTS)}) '
        f'AND status changed to Done during ("{start}", "{end}") '
        f'AND status != Discarded '
        f'AND issuetype not in (Epic)'
    )

    print("\n1. Fetching & filtering issues...")
    issues = get_all_issues(jql)

    # Double-filter: exclude Epics in case JQL filter missed any
    before = len(issues)
    issues = [i for i in issues
              if (i["fields"].get("issuetype") or {}).get("name", "").lower()
              not in EXCLUDED_ISSUE_TYPES]
    if before != len(issues):
        print(f"   ⚠️  Excluded {before - len(issues)} Epic(s) — {len(issues)} remaining")
    print(f"   ✅ {len(issues)} issues qualify (SP > 0 or Fix Version)")

    print("\n2. Fetching changelogs in parallel (10 threads)...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        futures    = {ex.submit(get_changelog, i["key"]): i for i in issues}
        changelogs = {}
        done_count = 0
        for future in concurrent.futures.as_completed(futures):
            issue = futures[future]
            changelogs[issue["key"]] = future.result()
            done_count += 1
            if done_count % 20 == 0:
                print(f"   {done_count}/{len(issues)} changelogs fetched...")
    print(f"   ✅ All {len(issues)} changelogs fetched")

    print("\n3. Computing cycle times (working days, PT holidays excluded)...")
    by_project  = defaultdict(list)
    all_tickets = []

    for issue in issues:
        key     = issue["key"]
        proj    = issue["fields"]["project"]["key"]
        summary = issue["fields"].get("summary", "")
        cl      = changelogs[key]

        cycle, blocked, in_progress, status_log = compute_cycle_time(cl["histories"])

        sp = (issue["fields"].get("customfield_10016")
           or issue["fields"].get("customfield_10004")
           or cl.get("story_points"))

        issuetype = (issue["fields"].get("issuetype") or {}).get("name", "").lower()
        fix_versions = (
            [v.get("name", "") for v in (issue["fields"].get("fixVersions") or [])]
            or cl.get("fix_versions", [])
        )

        entry = {
            "key":          key,
            "project":      proj,
            "summary":      summary,
            "cycle":        cycle,
            "blocked":      blocked,
            "in_progress":  in_progress,
            "status_log":   status_log,
            "story_points": sp,
            "fix_versions": fix_versions,
            "ac_failed":    cl["ac_failed"],
            "ai_tools":     cl["ai_tools"],
            "issuetype":    issuetype,
        }
        by_project[proj].append(entry)
        all_tickets.append(entry)

    print(f"\n{'─' * 60}")
    print(f"{'Project':<10} {'Tickets':>7}  {'AVG Cycle':>10}  {'AVG Blocked':>12}")
    print(f"{'─' * 60}")

    results   = []
    total_w   = total_n = 0

    for proj in PROJECTS:
        items = by_project.get(proj, [])
        avg_c = sum(i["cycle"]   for i in items) / len(items) if items else None
        avg_b = sum(i["blocked"] for i in items) / len(items) if items else 0.0
        if items and avg_c is not None:
            total_w += avg_c * len(items)
            total_n += len(items)

        sp_items    = [i for i in items if i.get("story_points") is not None]
        sp_coverage = len(sp_items) / len(items) if items else 0
        total_sp    = sum(float(i["story_points"]) for i in sp_items)
        tput_sp     = round(total_sp / weeks, 1) if weeks > 0 and sp_coverage >= 0.5 else None
        tput_sp_dev = round(total_sp / weeks / TEAM_DEVS.get(proj, 1), 1) if tput_sp is not None else None
        sp_na       = sp_coverage < 0.5

        REWORK_TYPES = {"story", "task", "sub-task"}
        rework_items = [i for i in items if i.get("issuetype", "") in REWORK_TYPES]
        rework_n     = sum(1 for i in rework_items if i.get("ac_failed"))
        rework_rate  = round(rework_n / len(rework_items) * 100) if rework_items else None

        ai_n    = sum(1 for i in items if i.get("ai_tools"))
        ai_rate = round(ai_n / len(items) * 100) if items else None

        total_cycle_wd  = sum(i["cycle"]       for i in items)
        total_inprog_wd = sum(i["in_progress"] for i in items)
        efficiency      = round(total_inprog_wd / total_cycle_wd * 100) if total_cycle_wd > 0 else None

        results.append({
            "project":           proj,
            "count":             len(items),
            "avg":               round(avg_c, 1) if avg_c is not None else None,
            "avg_blocked":       round(avg_b, 1),
            "throughput_sp":     tput_sp,
            "throughput_sp_dev": tput_sp_dev,
            "sp_coverage":       round(sp_coverage, 2),
            "total_sp":          total_sp,
            "sp_na":             sp_na,
            "rework_rate":       rework_rate,
            "ai_rate":           ai_rate,
            "efficiency":        efficiency,
        })
        bl_str = f"  (🚫 {avg_b:.1f}wd blocked)" if avg_b > 0 else ""
        print(f"{proj:<10} {len(items):>7}  {avg_c:>9.1f}wd{bl_str}" if avg_c is not None
              else f"{proj:<10} {len(items):>7}  {'—':>9}")

    overall = total_w / total_n if total_n else 0
    print(f"{'─' * 60}")
    print(f"{'OVERALL':<10} {total_n:>7}  {overall:>9.1f}wd")
    print(f"{'─' * 60}")
    print("\n✅ Working days only · SP > 0 or Fix Version · "
          "Ready, Blocked & Test Blocked excluded · PT holidays excluded")

    print("\n4. Fetching open bugs...")
    open_bugs  = get_open_bugs()
    print(f"   ✅ {open_bugs['total']} open bugs/defects found")
    for proj, n in open_bugs["by_project"].items():
        print(f"      {proj}: {n}")

    print("\n5. Generating Excel...")
    excel_file = export_excel(results, all_tickets, year, month, weeks)

    print("\n6. Uploading to Confluence...")
    confluence_url = upload_to_confluence(excel_file)

    return {
        "label":          label,
        "projects":       results,
        "total":          total_n,
        "overall":        round(overall, 1),
        "weeks":          round(weeks, 1),
        "confluence_url": confluence_url,
        "open_bugs":      open_bugs,
    }


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    today = date.today()
    year  = today.year if today.month > 1 else today.year - 1
    month = today.month - 1 or 12

    if len(sys.argv) == 3:
        year, month = int(sys.argv[1]), int(sys.argv[2])

    report = run_report(year, month)
    post_to_slack(report, confluence_url=report.get("confluence_url"))
